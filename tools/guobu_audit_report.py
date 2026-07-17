"""Business display helpers for the Guobu audit report."""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import uuid
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
try:
    from tools.guobu_audit_contract import network_failure
except ModuleNotFoundError:
    from guobu_audit_contract import network_failure


_UNKNOWN_REASON = "图片信息无法确认"

_STANDARD_REASONS = {
    "PRODUCT_TYPE_MISMATCH": "商品类型不一致",
    "PRODUCT_PHOTO_INVALID": "商品照片不符合要求",
    "UNBOXING_PHOTO_INVALID": "拆封/安装照片不符合要求",
    "ACTIVATION_PHOTO_INVALID": "激活照片不符合要求",
    "SN_MISSING_IN_ACTIVATION_PHOTO": "激活照片不符合要求",
    "ADDRESS_TOO_COARSE": "收货地址不符合要求",
    "DUPLICATE_IMAGE_EVIDENCE": "存在重复图片，不符合要求",
    "NON_REAL_PHOTO_REVIEW": "图片疑似非实拍",
    "NON_REAL_PHOTO_STRONG_RISK": "图片疑似非实拍",
    "IMAGE_STRONG_RISK": "图片疑似非实拍",
    "SN_MISMATCH": "SN不一致",
    "INVOICE_ORANGE_WARNING": "发票疑似已红冲",
    "MODEL_UNCERTAIN": _UNKNOWN_REASON,
    "PHOTO_AUTHENTICITY_SERVICE_FAILURE": "审核服务异常",
    "ARTIFACT_LOAD_FAILURE": "审核服务异常",
    "FFT_FAILURE": "审核服务异常",
    "SN_TRUNCATED_OBSCURED": "SN不完整，无法识别",
    "SN_NOT_FOUND": "SN无法识别",
    "SYSTEM_SN_MISSING": "系统SN缺失",
    "IMAGE_MISSING": "图片缺失",
    "FIELD_MISSING": "订单信息缺失",
    "PRODUCT_TYPE_MISSING": "商品类型信息缺失",
    "NON_REAL_PHOTO_FFT_RESCUE": "图片疑似非实拍",
}


def _indexed(items: list[dict], label: str) -> dict[str, dict]:
    result = {}
    for item in items:
        order_id = str(((item.get("row") or {}).get("id")) or "").strip()
        if not order_id:
            raise ValueError(f"empty {label} order ID")
        if order_id in result:
            raise ValueError(f"duplicate {label} order ID: {order_id!r}")
        result[order_id] = item
    return result


def _number(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) and not isinstance(value, bool) else 0.0


def _validated_elapsed(value: object, label: str) -> int | float:
    if type(value) not in (int, float) or not math.isfinite(value) or value < 0:
        raise ValueError(f"{label} elapsed must be a finite nonnegative number")
    return value


def _usage_objects(item: dict):
    roots = []
    root_ids = set()
    if isinstance(item.get("_raw"), dict):
        roots.append(item["_raw"])
        root_ids.add(id(item["_raw"]))
    row_raw = (item.get("row") or {}).get("_raw")
    if isinstance(row_raw, dict) and id(row_raw) not in root_ids:
        roots.append(row_raw)

    def visit(container: dict):
        for key, value in container.items():
            if key.endswith("_usage") and isinstance(value, dict):
                stage = key.split("_usage", 1)[0]
                yield value, container.get(f"{stage}_cached") is True
            elif isinstance(value, dict):
                yield from visit(value)

    seen_usage_ids = set()
    for root in roots:
        for usage, cached in visit(root):
            usage_id = id(usage)
            if usage_id not in seen_usage_ids:
                seen_usage_ids.add(usage_id)
                yield usage, cached


def _account_attempt(item: dict, source: str, order_timeout_seconds: float) -> tuple[dict, dict]:
    totals = {key: 0 for key in ("logical_input_tokens", "logical_cached_input_tokens",
        "logical_output_tokens", "billed_input_tokens", "billed_cached_input_tokens",
        "billed_output_tokens")}
    for usage, stage_cached in _usage_objects(item):
        prompt = int(_number(usage.get("prompt_tokens")))
        details = usage.get("prompt_tokens_details") or {}
        cached_input = min(prompt, int(_number(details.get("cached_tokens"))))
        output = int(_number(usage.get("completion_tokens")))
        totals["logical_input_tokens"] += prompt - cached_input
        totals["logical_cached_input_tokens"] += cached_input
        totals["logical_output_tokens"] += output
        if not stage_cached:
            totals["billed_input_tokens"] += prompt - cached_input
            totals["billed_cached_input_tokens"] += cached_input
            totals["billed_output_tokens"] += output
    row = item.get("row") or {}
    raw_elapsed = _validated_elapsed(row.get("elapsed_sec"), "attempt")
    effective_elapsed = (min(raw_elapsed, order_timeout_seconds)
                         if network_failure(item) else raw_elapsed)
    trace = {"source": source, "order_id": str(row.get("id") or "").strip(),
             "elapsed_seconds": raw_elapsed, "effective_elapsed_seconds": effective_elapsed,
             "item": item, **totals}
    return trace, totals


def _validate_final(item: dict) -> tuple[bool, str]:
    row = item.get("row") or {}
    manual = parse_manual_flag(row.get("manual_flag"))
    code = str(row.get("manual_reason_code") or "").strip()
    reason = str(row.get("manual_reason") or row.get("manual_reason_cn") or "").strip()
    if not manual and (code or reason):
        raise ValueError("manual_flag false contradicts final reason")
    if manual and not code:
        raise ValueError("manual_flag true requires final primary reason code")
    return manual, code


def _validate_flag_syntax(items: list[dict]) -> None:
    for item in items:
        parse_manual_flag((item.get("row") or {}).get("manual_flag"))


def merge_attempts(first_items: list[dict], retry_items: list[dict],
                   retry_ids: set[str] | None = None,
                   order_timeout_seconds: float = 60) -> tuple[list[dict], dict]:
    if (type(order_timeout_seconds) not in (int, float)
            or not math.isfinite(order_timeout_seconds) or order_timeout_seconds <= 0):
        raise ValueError("order timeout must be a positive finite number")
    first = _indexed(first_items, "first-run")
    retries = _indexed(retry_items, "retry")
    _validate_flag_syntax(first_items)
    _validate_flag_syntax(retry_items)
    failures = {order_id for order_id, item in first.items() if network_failure(item)}
    retry_item_ids = set(retries)
    if retry_ids is not None:
        selected = {str(order_id).strip() for order_id in retry_ids}
        if selected != retry_item_ids:
            raise ValueError("retry selection mismatch")
        if selected - failures:
            raise ValueError("retry ID is not a first-run network failure")
    if retry_item_ids - failures:
        raise ValueError(f"unknown retry order ID: {sorted(retry_item_ids - failures)!r}")

    accounting = {"attempts": [], "elapsed_seconds": 0.0, "raw_elapsed_seconds": 0.0,
                  "order_timeout_seconds": order_timeout_seconds,
                  **{key: 0 for key in ("logical_input_tokens", "logical_cached_input_tokens",
                     "logical_output_tokens", "billed_input_tokens", "billed_cached_input_tokens",
                     "billed_output_tokens")}}
    for source, items in (("first", first_items), ("retry", retry_items)):
        for item in items:
            trace, totals = _account_attempt(item, source, order_timeout_seconds)
            accounting["attempts"].append(trace)
            accounting["raw_elapsed_seconds"] += trace["elapsed_seconds"]
            accounting["elapsed_seconds"] += trace["effective_elapsed_seconds"]
            for key, value in totals.items():
                accounting[key] += value

    merged = []
    for order_id, first_item in first.items():
        retry_item = retries.get(order_id)
        final_item = retry_item or first_item
        manual, code = _validate_final(final_item)
        merged.append({"row": final_item.get("row") or {}, "reason_code": code,
                       "reason": standard_reason(code), "manual": manual,
                       "final_source": "retry" if retry_item else "first",
                       "first_attempt": first_item, "retry_attempt": retry_item})
    return merged, accounting


def _rate(numerator: int, denominator: int) -> dict:
    return {"numerator": numerator, "denominator": denominator,
            "rate": numerator / denominator if denominator else None}


def build_summary(rows: list[dict], accounting: dict, prices: dict | None = None) -> dict:
    failed_denominator = failed_numerator = passed_denominator = passed_numerator = 0
    manual_total = 0
    for item in rows:
        row = item.get("row") or {}
        manual = bool(item.get("manual"))
        manual_total += int(manual)
        status = str(row.get("source_flow_status") or "").strip()
        if status == "\u672a\u901a\u8fc7":
            failed_denominator += 1
            failed_numerator += int(manual)
        elif status == "\u5df2\u901a\u8fc7":
            passed_denominator += 1
            passed_numerator += int(manual)

    effective_hours = _number(accounting.get("elapsed_seconds")) / 3600
    sample_count = len(rows)
    human_rate = 550 / 7.5
    human_hours = sample_count / human_rate
    model_rate = sample_count / effective_hours if effective_hours else None
    multiple = model_rate / human_rate if model_rate is not None else None
    required = ("input_per_million", "cached_input_per_million", "output_per_million")
    configured = prices is not None and all(isinstance(prices.get(key), (int, float))
        and not isinstance(prices.get(key), bool) for key in required)
    cost = "\u5f85\u914d\u7f6e"
    if configured:
        cost = (accounting.get("billed_input_tokens", 0) * prices["input_per_million"]
                + accounting.get("billed_cached_input_tokens", 0) * prices["cached_input_per_million"]
                + accounting.get("billed_output_tokens", 0) * prices["output_per_million"]) / 1_000_000
    raw_elapsed = _number(accounting.get("raw_elapsed_seconds"))
    timeout = _number(accounting.get("order_timeout_seconds"))
    elapsed_note = ("\u6709\u6548\u5ba1\u6838\u603b\u7528\u65f6\u4e3a\u7d2f\u8ba1\u6bcf\u8ba2\u5355\u5904\u7406\u65f6\u957f\uff0c"
                    "\u5e76\u975e\u6279\u6b21\u5899\u949f\u541e\u5410\u91cf\u3002\u7f51\u7edc\u5931\u8d25\u5355\u6b21\u6709\u6548\u7528\u65f6\u6309 "
                    f"{timeout:g} \u79d2\u5c01\u9876\uff1b\u539f\u59cb\u7d2f\u8ba1 {raw_elapsed:.2f} \u79d2\uff0c"
                    f"\u6709\u6548\u7d2f\u8ba1 {_number(accounting.get('elapsed_seconds')):.2f} \u79d2\u3002")
    return {"sample_count": sample_count, "manual_total": manual_total,
            "automatic_pass_total": sample_count - manual_total,
            "failed_interception": _rate(failed_numerator, failed_denominator),
            "passed_false_positive": _rate(passed_numerator, passed_denominator),
            "billed_input_tokens": accounting.get("billed_input_tokens", 0),
            "billed_cached_input_tokens": accounting.get("billed_cached_input_tokens", 0),
            "billed_output_tokens": accounting.get("billed_output_tokens", 0),
            "estimated_cost": cost, "effective_hours": effective_hours,
            "human_orders_per_hour": human_rate, "human_estimated_hours": human_hours,
            "model_orders_per_hour": model_rate, "efficiency_multiple": multiple,
            "efficiency_improvement": multiple - 1 if multiple is not None else None,
            "saved_human_hours": human_hours - effective_hours,
            "raw_elapsed_seconds": raw_elapsed, "order_timeout_seconds": timeout,
            "elapsed_note": elapsed_note}


def standard_reason(code: str) -> str:
    """Return the approved business wording for a primary reason code."""
    if not code:
        return ""
    return _STANDARD_REASONS.get(code, _UNKNOWN_REASON)


def parse_manual_flag(value: object) -> bool:
    """Parse the only explicit values permitted for a final manual flag."""
    if type(value) is bool:
        return value
    if value == "是":
        return True
    if value == "否":
        return False
    raise ValueError(f"invalid manual_flag: {value!r}")


def _sn_value(value: object) -> str:
    return "" if value is None else str(value)


def _adjacent_transposition(system: str, observed: str) -> int | None:
    differences = [i for i, pair in enumerate(zip(system, observed)) if pair[0] != pair[1]]
    if len(differences) != 2:
        return None
    first, second = differences
    if second == first + 1 and system[first] == observed[second] and system[second] == observed[first]:
        return first
    return None


def _single_edit(longer: str, shorter: str) -> tuple[int, str] | None:
    for index in range(len(longer)):
        if longer[:index] + longer[index + 1 :] == shorter:
            return index, longer[index]
    return None


def _sn_difference(system: str, observed: str) -> str:
    if not observed:
        return "模型未读取到SN"
    if system == observed:
        return ""

    transposition = _adjacent_transposition(system, observed)
    if transposition is not None:
        end = transposition + 2
        return f"字符顺序不同：系统{system[transposition:end]}，模型{observed[transposition:end]}"

    if len(system) == len(observed):
        differences = [i for i, pair in enumerate(zip(system, observed)) if pair[0] != pair[1]]
        if len(differences) == 1:
            index = differences[0]
            return f"第{index + 1}位不同：系统{system[index]}，模型{observed[index]}"
        return "SN存在多处差异"

    if system.startswith(observed):
        return f"模型末尾少读{system[len(observed):]}"
    if system.endswith(observed):
        return f"模型开头少读{system[:-len(observed)]}"

    if len(observed) == len(system) + 1:
        edit = _single_edit(observed, system)
        if edit is not None:
            index, character = edit
            return f"模型第{index + 1}位多读{character}"
    if len(system) == len(observed) + 1:
        edit = _single_edit(system, observed)
        if edit is not None:
            index, character = edit
            return f"模型第{index + 1}位少读{character}"
    return "SN存在多处差异"


def sn_display(row: dict) -> tuple[str, str]:
    """Return the fixed status and raw-character difference for an audit row."""
    system = _sn_value(row.get("system_sn"))
    observed = _sn_value(row.get("observed_sn"))
    if row.get("sn_match") is True:
        return "是", ""
    if not system:
        return "无系统SN", ""
    if not observed:
        return "未读取", _sn_difference(system, observed)
    return "否", _sn_difference(system, observed)


_DETAIL_HEADERS = ["订单号", "是否转人工", "原始流程状态", "转人工原因",
                   "系统SN", "模型SN", "SN是否一致", "SN具体差别"]
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _safe_text(value: object) -> str:
    return "" if value is None else str(value)


def _validate_report_rows(rows: list[dict]) -> None:
    if not rows:
        raise ValueError("empty rows")
    seen = set()
    for item in rows:
        if not isinstance(item, dict) or not isinstance(item.get("row"), dict):
            raise ValueError("malformed merged row")
        row = item["row"]
        order_id = str(row.get("id") or "").strip()
        if (not order_id or type(item.get("manual")) is not bool
                or not isinstance(item.get("reason_code"), str)
                or item.get("final_source") not in {"first", "retry"}):
            raise ValueError("malformed merged row")
        manual, code = _validate_final({"row": row})
        if item["manual"] is not manual or item["reason_code"].strip() != code:
            raise ValueError("malformed merged row business decision")
        if order_id in seen:
            raise ValueError(f"duplicate display order ID: {order_id!r}")
        seen.add(order_id)


def _validate_audit_json(audit_json: dict) -> None:
    if not isinstance(audit_json, dict):
        raise ValueError("audit JSON must be an object")
    accounting = audit_json.get("accounting")
    required_totals = ("elapsed_seconds", "raw_elapsed_seconds", "order_timeout_seconds",
                       "logical_input_tokens", "logical_cached_input_tokens",
                       "logical_output_tokens", "billed_input_tokens",
                       "billed_cached_input_tokens", "billed_output_tokens")
    if (not isinstance(accounting, dict) or not isinstance(accounting.get("attempts"), list)
            or not accounting["attempts"]
            or any(not isinstance(accounting.get(key), (int, float)) for key in required_totals)):
        raise ValueError("audit accounting and attempt trace are required")
    timeout = accounting["order_timeout_seconds"]
    if type(timeout) not in (int, float) or not math.isfinite(timeout) or timeout <= 0:
        raise ValueError("audit order timeout must be a positive finite number")
    _validated_elapsed(accounting["raw_elapsed_seconds"], "audit raw total")
    _validated_elapsed(accounting["elapsed_seconds"], "audit effective total")
    attempt_keys = {"source", "order_id", "elapsed_seconds", "effective_elapsed_seconds", "item"}
    for attempt in accounting["attempts"]:
        if (not isinstance(attempt, dict) or not attempt_keys <= attempt.keys()
                or attempt["source"] not in {"first", "retry"}
                or not str(attempt["order_id"]).strip()
                or not isinstance(attempt["item"], dict)):
            raise ValueError("audit attempt trace is malformed")
        _validated_elapsed(attempt["elapsed_seconds"], "audit attempt raw")
        _validated_elapsed(attempt["effective_elapsed_seconds"], "audit attempt effective")
    if "pricing" not in audit_json:
        raise ValueError("audit pricing assumption is required")
    pricing = audit_json["pricing"]
    price_keys = ("input_per_million", "cached_input_per_million", "output_per_million")
    if pricing is not None and (not isinstance(pricing, dict)
            or any(not isinstance(pricing.get(key), (int, float))
                   or isinstance(pricing.get(key), bool) or pricing[key] < 0 for key in price_keys)):
        raise ValueError("audit pricing is malformed")


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_summary(sheet, summary: dict) -> None:
    unavailable = "\u65e0\u53ef\u8ba1\u7b97\u6837\u672c"
    sheet.append(["\u6307\u6807", "\u503c"])
    rows = [
        ("\u6837\u672c\u603b\u6570", summary["sample_count"], "0"),
        ("\u8f6c\u4eba\u5de5\u603b\u6570", summary["manual_total"], "0"),
        ("\u81ea\u52a8\u901a\u8fc7\u603b\u6570", summary["automatic_pass_total"], "0"),
        ("\u672a\u901a\u8fc7\u62e6\u622a\u6570\uff08\u5206\u5b50\uff09", summary["failed_interception"]["numerator"], "0"),
        ("\u672a\u901a\u8fc7\u8ba2\u5355\u6570\uff08\u5206\u6bcd\uff09", summary["failed_interception"]["denominator"], "0"),
        ("\u672a\u901a\u8fc7\u62e6\u622a\u7387", summary["failed_interception"]["rate"] if summary["failed_interception"]["rate"] is not None else unavailable, "0.0%"),
        ("\u5df2\u901a\u8fc7\u8bef\u5224\u6570\uff08\u5206\u5b50\uff09", summary["passed_false_positive"]["numerator"], "0"),
        ("\u5df2\u901a\u8fc7\u8ba2\u5355\u6570\uff08\u5206\u6bcd\uff09", summary["passed_false_positive"]["denominator"], "0"),
        ("\u5df2\u901a\u8fc7\u8bef\u5224\u7387", summary["passed_false_positive"]["rate"] if summary["passed_false_positive"]["rate"] is not None else unavailable, "0.0%"),
        ("\u8f93\u5165Token", summary["billed_input_tokens"] + summary["billed_cached_input_tokens"], "0"),
        ("\u8f93\u51faToken", summary["billed_output_tokens"], "0"),
        ("Token\u603b\u6d88\u8017", summary["billed_input_tokens"] + summary["billed_cached_input_tokens"] + summary["billed_output_tokens"], "0"),
        ("Token\u9884\u8ba1\u6210\u672c", summary["estimated_cost"], "0.00"),
        ("\u6709\u6548\u5ba1\u6838\u603b\u7528\u65f6\uff08\u5c0f\u65f6\uff09", summary["effective_hours"], "0.00"),
        ("\u4eba\u5de5\u6bcf\u5c0f\u65f6\u5ba1\u6838\u91cf", summary["human_orders_per_hour"], "0.00"),
        ("\u4eba\u5de5\u9884\u8ba1\u7528\u65f6\uff08\u5c0f\u65f6\uff09", summary["human_estimated_hours"], "0.00"),
        ("\u6a21\u578b\u6bcf\u5c0f\u65f6\u5ba1\u6838\u91cf", summary["model_orders_per_hour"] if summary["model_orders_per_hour"] is not None else unavailable, "0.00"),
        ("\u6548\u7387\u500d\u6570", summary["efficiency_multiple"] if summary["efficiency_multiple"] is not None else unavailable, "0.0x"),
        ("\u6548\u7387\u63d0\u5347\u7387", summary["efficiency_improvement"] if summary["efficiency_improvement"] is not None else unavailable, "0.0%"),
        ("\u9884\u8ba1\u8282\u7701\u4eba\u5de5\u65f6\u95f4\uff08\u5c0f\u65f6\uff09", summary["saved_human_hours"], "0.00"),
        ("\u53e3\u5f84\u8bf4\u660e", summary["elapsed_note"] + " \u516c\u5f0f\u5b9a\u4e49\u4fdd\u7559\u5728 JSON trace\u3002", "General"),
    ]
    for label, value, number_format in rows:
        sheet.append([label, value])
        sheet.cell(sheet.max_row, 2).number_format = number_format
    _style_header(sheet)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 62
    sheet["B22"].alignment = Alignment(wrap_text=True)


def write_report(rows: list[dict], summary: dict, audit_json: dict,
                 xlsx_path: str | Path, json_path: str | Path) -> None:
    """Write the compact business workbook and its traceable UTF-8 audit JSON."""
    _validate_report_rows(rows)
    _validate_audit_json(audit_json)
    xlsx_path, json_path = Path(xlsx_path), Path(json_path)
    if xlsx_path.resolve() == json_path.resolve():
        raise ValueError("XLSX and JSON output paths must be distinct")
    if xlsx_path.exists() or json_path.exists():
        raise FileExistsError("report output already exists")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    detail.append(_DETAIL_HEADERS)
    for item in rows:
        row = item["row"]
        status, difference = sn_display(row)
        detail.append([str(row["id"]), "是" if item["manual"] else "否",
                       _safe_text(row.get("source_flow_status")), standard_reason(item["reason_code"]),
                       _safe_text(row.get("system_sn")), _safe_text(row.get("observed_sn")),
                       status, _safe_text(difference)])
        for column in (1, 3, 4, 5, 6, 8):
            detail.cell(detail.max_row, column).data_type = "s"
        for column in (1, 5, 6):
            detail.cell(detail.max_row, column).number_format = "@"
        for column in (4, 8):
            detail.cell(detail.max_row, column).alignment = Alignment(wrap_text=True, vertical="top")
    _style_header(detail)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:H{detail.max_row}"
    table = Table(displayName="明细表格", ref=detail.auto_filter.ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                          showFirstColumn=False, showLastColumn=False)
    detail.add_table(table)
    for column, width in zip("ABCDEFGH", (24, 14, 18, 34, 24, 24, 16, 38)):
        detail.column_dimensions[column].width = width

    summary_sheet = workbook.create_sheet("汇总表")
    _add_summary(summary_sheet, summary)
    for sheet in workbook:
        for row in sheet:
            for cell in row:
                font = copy(cell.font)
                font.name = "Arial"
                cell.font = font
    workbook.save(xlsx_path)
    payload = dict(audit_json)
    payload["summary"] = summary
    payload["rows"] = rows
    payload["summary_formula_definitions"] = {
        "\u672a\u901a\u8fc7\u62e6\u622a\u7387": '=IF(B6=0,"\u65e0\u53ef\u8ba1\u7b97\u6837\u672c",B5/B6)',
        "\u5df2\u901a\u8fc7\u8bef\u5224\u7387": '=IF(B9=0,"\u65e0\u53ef\u8ba1\u7b97\u6837\u672c",B8/B9)',
        "\u6a21\u578b\u6bcf\u5c0f\u65f6\u5ba1\u6838\u91cf": '=IF(B15=0,"\u65e0\u53ef\u8ba1\u7b97\u6837\u672c",B2/B15)',
        "\u6548\u7387\u500d\u6570": '=IF(B16=0,"\u65e0\u53ef\u8ba1\u7b97\u6837\u672c",B18/B16)',
        "\u6548\u7387\u63d0\u5347\u7387": '=IF(B19="\u65e0\u53ef\u8ba1\u7b97\u6837\u672c","\u65e0\u53ef\u8ba1\u7b97\u6837\u672c",B19-1)',
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_jsonl(path: str | None) -> list[dict]:
    if not path:
        return []
    return [json.loads(line) for line in Path(path).read_text(encoding="utf-8").splitlines()
            if line.strip()]


def _read_jsonl_files(paths: list[str], label: str) -> list[dict]:
    combined = []
    for path in paths:
        items = _read_jsonl(path)
        if not items:
            raise ValueError(f"empty {label} JSONL: {path}")
        combined.extend(items)
    return combined


def _retry_ids(path: str | None) -> set[str] | None:
    if not path:
        return None
    value = json.loads(Path(path).read_text(encoding="utf-8"))
    metadata = value if isinstance(value, dict) else None
    if metadata is not None:
        orders = value.get("orders")
        missing = value.get("missing")
        requested = value.get("requested")
        selected = value.get("selected")
        if not isinstance(orders, list):
            raise ValueError("retry selection orders must be a JSON list")
        if not isinstance(missing, list) or missing:
            raise ValueError("retry selection missing orders must be an empty JSON list")
        value = orders
    if not isinstance(value, list):
        raise ValueError("retry selection must be a JSON list")
    normalized = set()
    for item in value:
        if type(item) not in (str, int, float):
            raise ValueError("retry selection order ID must be a string or number")
        order_id = str(item).strip()
        if not order_id:
            raise ValueError("retry selection order ID must not be empty")
        if order_id in normalized:
            raise ValueError(f"duplicate retry selection order ID: {order_id!r}")
        normalized.add(order_id)
    if metadata is not None:
        counts = (requested, selected)
        if any(type(count) not in (int, float) or count < 0 or int(count) != count
               for count in counts):
            raise ValueError("retry selection requested and selected must be integer counts")
        if int(requested) != int(selected) or int(selected) != len(normalized):
            raise ValueError("retry selection requested/selected count mismatch")
    if not normalized:
        raise ValueError("retry selection file must not be empty")
    return normalized


def _retry_ids_many(paths: list[str]) -> set[str] | None:
    if not paths:
        return None
    combined = set()
    for path in paths:
        ids = _retry_ids(path)
        duplicates = combined & ids
        if duplicates:
            raise ValueError(f"duplicate retry selection order ID across files: {sorted(duplicates)!r}")
        combined.update(ids)
    return combined


def _replace_outputs(temp_outputs: list[Path], outputs: list[Path]) -> None:
    backups: list[Path | None] = []
    replaced = []
    try:
        for output in outputs:
            if output.exists():
                backup = output.with_name(f".{output.name}.{uuid.uuid4().hex}.bak")
                shutil.copy2(output, backup)
                backups.append(backup)
            else:
                backups.append(None)
        for temp, output in zip(temp_outputs, outputs):
            os.replace(temp, output)
            replaced.append(output)
    except Exception:
        for output, backup in zip(outputs, backups):
            if backup is not None and backup.exists():
                os.replace(backup, output)
            elif output in replaced and output.exists():
                output.unlink()
        raise
    finally:
        for path in [*temp_outputs, *(backup for backup in backups if backup is not None)]:
            if path.exists():
                path.unlink()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate a Guobu audit XLSX and trace JSON")
    parser.add_argument("--first-jsonl", required=True)
    parser.add_argument("--retry-jsonl", action="append", default=[])
    parser.add_argument("--retry-selection-json", action="append", default=[])
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--output-json", required=True)
    parser.add_argument("--input-price-per-million", type=float)
    parser.add_argument("--cached-input-price-per-million", type=float)
    parser.add_argument("--output-price-per-million", type=float)
    parser.add_argument("--order-timeout-seconds", type=float, default=60)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args(argv)
    outputs = [Path(args.output_xlsx), Path(args.output_json)]
    if outputs[0].resolve() == outputs[1].resolve():
        raise ValueError("XLSX and JSON output paths must be distinct")
    if not args.overwrite and any(output.exists() for output in outputs):
        raise FileExistsError("report output already exists")
    prices = None
    supplied = [args.input_price_per_million, args.cached_input_price_per_million,
                args.output_price_per_million]
    if any(value is not None for value in supplied):
        if not all(value is not None and value >= 0 for value in supplied):
            parser.error("all three non-negative prices are required")
        prices = {"input_per_million": supplied[0], "cached_input_per_million": supplied[1],
                  "output_per_million": supplied[2]}
    rows, accounting = merge_attempts(_read_jsonl(args.first_jsonl),
                                      _read_jsonl_files(args.retry_jsonl, "retry"),
                                      _retry_ids_many(args.retry_selection_json),
                                      args.order_timeout_seconds)
    summary = build_summary(rows, accounting, prices)
    audit_json = {"summary": summary, "accounting": accounting, "pricing": prices,
                  "sources": {"first_jsonl": args.first_jsonl,
                              "retry_jsonl": args.retry_jsonl,
                              "retry_selection_json": args.retry_selection_json},
                  "rows": rows}
    if args.overwrite:
        token = uuid.uuid4().hex
        temp_outputs = [outputs[0].with_name(f".{outputs[0].stem}.{token}.tmp.xlsx"),
                        outputs[1].with_name(f".{outputs[1].stem}.{token}.tmp.json")]
        try:
            write_report(rows, summary, audit_json, temp_outputs[0], temp_outputs[1])
            load_workbook(temp_outputs[0], read_only=True).close()
            json.loads(temp_outputs[1].read_text(encoding="utf-8"))
            _replace_outputs(temp_outputs, outputs)
        finally:
            for temp in temp_outputs:
                if temp.exists():
                    temp.unlink()
    else:
        write_report(rows, summary, audit_json, outputs[0], outputs[1])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
