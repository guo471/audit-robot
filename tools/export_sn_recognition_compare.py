# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import os
import re
import sys
import time
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.run_guobu_model_audit_v2 import _with_detail, normalize_sn, precheck_task


DIRECT_SN_PROMPT = """只读取图片里的 SN / S/N / Serial Number / 序列号。
只输出一个结果，不要解释，不要标点，不要 JSON，不要 Markdown。
如果能看到完整 SN，只输出 SN 原文。
如果看不到完整 SN，只输出 SN_NOT_FOUND。
不要猜测、纠正或补全；O 和 0、I 和 1、L 和 1、S 和 5、B 和 8 必须按图片原样输出。"""


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _task_paths(tasks_dir: Path, limit: int | None) -> list[Path]:
    paths = sorted(tasks_dir.glob("*.json"))
    return paths[:limit] if limit else paths


def _image_url(image: dict[str, Any]) -> str:
    url = image.get("source_url") or image.get("url")
    if url:
        return str(url)
    local_path = image.get("local_path")
    if not local_path:
        return ""
    path = Path(str(local_path))
    if not path.exists():
        return ""
    mime = mimetypes.guess_type(path.name)[0] or "image/jpeg"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _clean_sn_text(text: str) -> str:
    value = (text or "").strip()
    value = re.sub(r"^```(?:text)?|```$", "", value, flags=re.IGNORECASE).strip()
    value = value.strip("\"'`：: ，,。;；")
    if not value:
        return ""
    if "SN_NOT_FOUND" in value.upper():
        return "SN_NOT_FOUND"
    value = re.sub(r"^(?:SN|S/N|SERIAL\s*(?:NO\.?|NUMBER)?|序列号)\s*[:：]?\s*", "", value, flags=re.IGNORECASE).strip()
    parts = re.findall(r"[A-Za-z0-9][A-Za-z0-9_./: -]{4,}[A-Za-z0-9]", value)
    if parts:
        return max(parts, key=len).strip()
    return value.splitlines()[0].strip()


def _call_direct_sn(base_url: str, api_key: str, model: str, images: list[dict[str, Any]]) -> tuple[str, float]:
    content: list[dict[str, Any]] = [{"type": "text", "text": "只输出图片中最完整的 SN。"}]
    for image in images:
        url = _image_url(image)
        if url:
            content.append({"type": "image_url", "image_url": {"url": url, "detail": image.get("_detail") or "high"}})
    body: dict[str, Any] = {
        "model": model,
        "messages": [{"role": "system", "content": DIRECT_SN_PROMPT}, {"role": "user", "content": content}],
    }
    normalized_model = (model or "").strip().lower()
    if normalized_model.startswith("qwen3.") or normalized_model.startswith("qwen-"):
        body["enable_thinking"] = False
    request = urllib.request.Request(
        base_url.rstrip("/") + "/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": "Bearer " + api_key, "Content-Type": "application/json"},
    )
    started = time.time()
    with urllib.request.urlopen(request, timeout=40) as response:
        raw = response.read().decode("utf-8")
    elapsed = time.time() - started
    response_data = json.loads(raw)
    return _clean_sn_text(response_data["choices"][0]["message"]["content"]), elapsed


def _recognize_one(base_url: str, api_key: str, model: str, task_path: Path, cache_dir: Path) -> dict[str, Any]:
    task = json.loads(task_path.read_text(encoding="utf-8-sig"))
    fields = task.get("fields") or {}
    order_id = _as_text(task.get("channel_order_no") or fields.get("channel_order_no") or task_path.stem)
    system_sn = _as_text(fields.get("system_sn"))
    started = time.time()
    row: dict[str, Any] = {
        "订单号": order_id,
        "系统SN": system_sn,
        "模型识别SN": "",
        "模型归一SN": "",
        "是否一致": "",
        "耗时秒": "",
        "错误": "",
    }
    try:
        precheck = precheck_task(task)
        activation_images = precheck.get("activation_images") or []
        images = _with_detail(activation_images, "high")
        observed, elapsed = _call_direct_sn(base_url, api_key, model, images)
        normalized_observed = normalize_sn(observed) if observed != "SN_NOT_FOUND" else ""
        normalized_system = normalize_sn(system_sn)
        row.update(
            {
                "模型识别SN": observed,
                "模型归一SN": normalized_observed,
                "是否一致": "是" if normalized_system and normalized_observed == normalized_system else "否",
                "耗时秒": round(elapsed, 2),
            }
        )
    except Exception as exc:
        row["耗时秒"] = round(time.time() - started, 2)
        row["是否一致"] = "错误"
        row["错误"] = f"{type(exc).__name__}: {exc}"
    return row


def _write_xlsx(rows: list[dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SN识别对比"
    headers = ["订单号", "系统SN", "模型识别SN", "模型归一SN", "是否一致", "耗时秒", "错误"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status = row[4].value
        if status == "是":
            row[4].fill = PatternFill("solid", fgColor="E2F0D9")
        elif status == "否":
            row[4].fill = PatternFill("solid", fgColor="FCE4D6")
        elif status == "错误":
            row[4].fill = PatternFill("solid", fgColor="FFF2CC")

    widths = [26, 24, 26, 26, 10, 10, 45]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tasks-dir", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--cache-dir", required=True)
    parser.add_argument("--model", default="qwen3.7-plus")
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    base_url = os.environ["VISION_API_BASE_URL"]
    api_key = os.environ["VISION_API_KEY"]
    paths = _task_paths(Path(args.tasks_dir), args.limit or None)
    rows: list[dict[str, Any]] = []
    for index, task_path in enumerate(paths, 1):
        print(f"[{index}/{len(paths)}] {task_path.stem}", flush=True)
        row = _recognize_one(base_url, api_key, args.model, task_path, Path(args.cache_dir))
        rows.append(row)
        print(f" -> {row['是否一致']} {row['模型识别SN']} {row['耗时秒']}s", flush=True)
    _write_xlsx(rows, Path(args.out))
    print(f"XLSX={args.out}")


if __name__ == "__main__":
    main()
