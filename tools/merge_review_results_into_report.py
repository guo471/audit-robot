import argparse
import json
import shutil
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def rows_by_order(ws):
    headers = [cell.value for cell in ws[1]]
    order_col = headers.index("订单号") + 1
    return headers, {str(ws.cell(row, order_col).value): row for row in range(2, ws.max_row + 1)}


def yn(value):
    return value == "是"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("main_report", type=Path)
    parser.add_argument("review_report", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--summary-json", type=Path)
    args = parser.parse_args()

    backup = args.main_report.with_name(args.main_report.stem + "_before_review_merge" + args.main_report.suffix)
    if not backup.exists():
        shutil.copy2(args.main_report, backup)
    if args.summary_json and args.summary_json.exists():
        json_backup = args.summary_json.with_name(args.summary_json.stem + "_before_review_merge.json")
        if not json_backup.exists():
            shutil.copy2(args.summary_json, json_backup)

    main_wb = load_workbook(args.main_report)
    review_wb = load_workbook(args.review_report, data_only=False)
    main_ws = main_wb.worksheets[0]
    review_ws = review_wb.worksheets[0]
    main_headers, main_orders = rows_by_order(main_ws)
    review_headers, review_orders = rows_by_order(review_ws)
    if main_headers != review_headers:
        raise SystemExit("detail headers do not match")

    missing = sorted(set(review_orders) - set(main_orders))
    if missing:
        raise SystemExit("missing order IDs: " + ", ".join(missing))
    if len(review_orders) != review_ws.max_row - 1:
        raise SystemExit("duplicate order IDs in review report")

    for order_id, review_row in review_orders.items():
        main_row = main_orders[order_id]
        for col in range(2, main_ws.max_column + 1):
            main_ws.cell(main_row, col).value = review_ws.cell(review_row, col).value

    detail = list(main_ws.iter_rows(min_row=2, values_only=True))
    manual = sum(yn(row[2]) for row in detail)
    first_timeout = sum(yn(row[13]) for row in detail)
    second_rerun = sum(row[17] is not None or row[20] is not None for row in detail)
    second_still_timeout = sum(yn(row[14]) for row in detail)
    second_resolved = second_rerun - second_still_timeout
    final_timeout = second_still_timeout
    final_tokens = sum((row[11] or 0) for row in detail)
    actual_tokens = sum((row[21] if row[21] is not None else row[11] or 0) for row in detail)
    final_elapsed = round(sum((row[10] or 0) for row in detail), 2)
    actual_elapsed = round(sum((row[18] if row[18] is not None else row[10] or 0) for row in detail), 2)
    reasons = Counter(row[3] if yn(row[2]) else "自动通过" for row in detail)

    summary = {
        "total": len(detail),
        "first_timeout": first_timeout,
        "second_rerun": second_rerun,
        "second_resolved": second_resolved,
        "second_still_timeout": second_still_timeout,
        "final_manual": manual,
        "final_auto": len(detail) - manual,
        "final_timeout": final_timeout,
        "final_tokens": final_tokens,
        "actual_tokens": actual_tokens,
        "final_elapsed_seconds": final_elapsed,
        "actual_elapsed_seconds": actual_elapsed,
    }

    summary_ws = main_wb.worksheets[1]
    template_styles = []
    for row in range(1, summary_ws.max_row + 1):
        template_styles.append([copy(summary_ws.cell(row, col)._style) for col in range(1, 3)])
    summary_ws.delete_rows(1, summary_ws.max_row)
    summary_rows = [
        ("指标", "数值"),
        ("样本总数", summary["total"]),
        ("第一次网络失败转人工单数", first_timeout),
        ("第二轮重跑单数", second_rerun),
        ("第二轮已解决网络失败单数", second_resolved),
        ("第二轮仍网络失败单数", second_still_timeout),
        ("综合后转人工总数", manual),
        ("综合后自动通过数", len(detail) - manual),
        ("综合后网络失败转人工数", final_timeout),
        ("综合结果token合计", final_tokens),
        ("两轮实际token合计", actual_tokens),
        ("综合结果耗时合计(秒)", final_elapsed),
        ("两轮实际耗时合计(秒)", actual_elapsed),
        (None, None),
        ("综合后转人工原因分布", "单数"),
    ]
    ordered_reasons = (["自动通过"] if "自动通过" in reasons else []) + [
        key for key, _ in reasons.most_common() if key != "自动通过"
    ]
    summary_rows.extend((key, reasons[key]) for key in ordered_reasons)
    for row_index, values in enumerate(summary_rows, 1):
        for col_index, value in enumerate(values, 1):
            cell = summary_ws.cell(row_index, col_index, value)
            style_row = min(row_index, len(template_styles)) - 1
            if style_row >= 0:
                cell._style = copy(template_styles[style_row][col_index - 1])

    output = args.output or args.main_report
    main_wb.save(output)
    if args.summary_json:
        args.summary_json.write_text(
            json.dumps({"summary": summary, "reason_counts": dict(reasons)}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    print(json.dumps({"merged": len(review_orders), "output": str(output), "backup": str(backup), "summary": summary, "reason_counts": dict(reasons)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
