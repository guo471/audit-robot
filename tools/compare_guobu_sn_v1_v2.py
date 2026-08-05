# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any, Iterable


HEADERS = [
    "订单号",
    "V1数据状态",
    "V1是否转人工",
    "V1原因码",
    "V1模型SN",
    "V1_SN是否一致",
    "V2数据状态",
    "V2是否转人工",
    "V2原因码",
    "V2模型SN",
    "V2_SN是否一致",
    "V2品类",
    "结论是否变化",
]

READABLE_HEADERS = [
    "渠道订单号",
    "结论是否一致",
    "新版SN",
    "旧版SN",
    "系统SN",
    "新版结果",
    "旧版结果",
    "新版原因码",
    "旧版原因码",
]

REASON_LABELS = {
    "PASS": "PASS（SN一致，通过）",
    "SN_ONLY_MATCH_NOT_FULL_AUDIT": "SN_ONLY_MATCH_NOT_FULL_AUDIT（SN一致，SN单项通过；非完整审核结论）",
    "SN_MISMATCH": "SN_MISMATCH（SN不一致）",
    "SN_NOT_FOUND": "SN_NOT_FOUND（未识别到完整可信SN）",
    "MODEL_UNCERTAIN": "MODEL_UNCERTAIN（模型不确定，转人工）",
    "SYSTEM_SN_MISSING": "SYSTEM_SN_MISSING（系统SN缺失）",
}


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig") as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number} is not a JSON object")
            entries.append(value)
    return entries


def _row(entry: dict[str, Any]) -> dict[str, Any]:
    value = entry.get("row") or entry.get("result") or entry
    return value if isinstance(value, dict) else {}


def _clean_identifier(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"[\u200B-\u200D\uFEFF]", "", text)
    text = re.sub(r"\s+", "", text)
    return text.lstrip("'")


def _order_id(entry: dict[str, Any]) -> str:
    row = _row(entry)
    task = entry.get("task") if isinstance(entry.get("task"), dict) else {}
    row_id = _clean_identifier(row.get("id") or "")
    task_order_id = _clean_identifier(task.get("channel_order_no") or "")
    task_id = str(task.get("task_id") or "")
    entry_order_id = _clean_identifier(entry.get("channel_order_no") or "")
    if row_id and task_order_id and row_id != task_order_id:
        raise ValueError(f"conflicting order IDs: row={row_id}, task={task_order_id}")
    if row_id and not task_order_id and task_id:
        equivalent_task_id = task_id == row_id or task_id.endswith(f"-{row_id}")
        if not equivalent_task_id:
            raise ValueError(f"conflicting order IDs: row={row_id}, task={task_id}")
    return row_id or task_order_id or task_id or entry_order_id


def _manual_flag(row: dict[str, Any]) -> str:
    if row.get("manual_flag") in {"是", "否"}:
        return str(row.get("manual_flag"))
    return "是" if row.get("manual_required") else "否"


def _sn_match_text(row: dict[str, Any]) -> str:
    value = row.get("sn_match")
    if value is True:
        return "是"
    if value is False:
        return "否"
    return ""


def _is_yes(value: Any) -> bool:
    return str(value or "").strip() == "是"


def _readable_result(value: Any) -> str:
    return "通过" if _is_yes(value) else "不通过/转人工"


def _readable_reason(raw_code: Any, passed: bool) -> str:
    code = str(raw_code or "").strip()
    if not code and passed:
        code = "PASS"
    if not code and not passed:
        code = "SN_MISMATCH"
    return REASON_LABELS.get(code, f"{code}（未登记中文说明）")


def _decision_signature(row: dict[str, Any]) -> tuple[str, str, str]:
    sn_match = _sn_match_text(row)
    reason_code = str(row.get("manual_reason_code") or "")
    if sn_match == "是" and reason_code == "SN_ONLY_MATCH_NOT_FULL_AUDIT":
        reason_code = ""
    observed_sn = re.sub(r"[^0-9A-Z]", "", str(row.get("observed_sn") or "").upper())
    return (
        sn_match,
        reason_code,
        observed_sn,
    )


def _index(entries: Iterable[dict[str, Any]], label: str) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for entry in entries:
        order_id = _order_id(entry)
        if not order_id:
            raise ValueError(f"{label} contains a row without order ID")
        if order_id in indexed:
            raise ValueError(f"{label} contains duplicate order ID: {order_id}")
        indexed[order_id] = entry
    return indexed


def compare_result_sets(
    v1_entries: Iterable[dict[str, Any]],
    v2_entries: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    v1 = _index(v1_entries, "V1")
    v2 = _index(v2_entries, "V2")
    rows: list[dict[str, Any]] = []
    for order_id in sorted(set(v1) | set(v2)):
        v1_row = _row(v1[order_id]) if order_id in v1 else {}
        v2_row = _row(v2[order_id]) if order_id in v2 else {}
        both = bool(v1_row and v2_row)
        changed = "无法比较" if not both else ("是" if _decision_signature(v1_row) != _decision_signature(v2_row) else "否")
        rows.append(
            {
                "订单号": order_id,
                "V1数据状态": "存在" if v1_row else "缺失",
                "V1是否转人工": _manual_flag(v1_row) if v1_row else "",
                "V1原因码": str(v1_row.get("manual_reason_code") or ""),
                "V1模型SN": str(v1_row.get("observed_sn") or ""),
                "V1_SN是否一致": _sn_match_text(v1_row),
                "V2数据状态": "存在" if v2_row else "缺失",
                "V2是否转人工": _manual_flag(v2_row) if v2_row else "",
                "V2原因码": str(v2_row.get("manual_reason_code") or ""),
                "V2模型SN": str(v2_row.get("observed_sn") or ""),
                "V2_SN是否一致": _sn_match_text(v2_row),
                "V2品类": str(v2_row.get("audit_category") or ""),
                "结论是否变化": changed,
            }
        )
    return rows


def write_outputs(rows: list[dict[str, Any]], output_stem: Path) -> tuple[Path, Path]:
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_stem.with_suffix(".csv")
    json_path = output_stem.with_suffix(".json")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return csv_path, json_path


def _system_sn_index(dataset_entries: Iterable[dict[str, Any]]) -> dict[str, str]:
    indexed: dict[str, str] = {}
    for entry in dataset_entries:
        order_id = _clean_identifier(entry.get("channel_order_no", ""))
        if order_id:
            indexed[order_id] = str(entry.get("system_sn") or "")
    return indexed


def build_readable_rows(
    rows: Iterable[dict[str, Any]],
    dataset_entries: Iterable[dict[str, Any]],
) -> list[dict[str, str]]:
    system_sn_by_order = _system_sn_index(dataset_entries)
    readable: list[dict[str, str]] = []
    for row in rows:
        order_id = _clean_identifier(row.get("订单号", ""))
        old_pass = _is_yes(row.get("V1_SN是否一致"))
        new_pass = _is_yes(row.get("V2_SN是否一致"))
        readable.append(
            {
                "渠道订单号": order_id,
                "结论是否一致": "一致" if old_pass == new_pass else "不一致",
                "新版SN": str(row.get("V2模型SN") or ""),
                "旧版SN": str(row.get("V1模型SN") or ""),
                "系统SN": system_sn_by_order.get(order_id, ""),
                "新版结果": _readable_result(row.get("V2_SN是否一致")),
                "旧版结果": _readable_result(row.get("V1_SN是否一致")),
                "新版原因码": _readable_reason(row.get("V2原因码"), new_pass),
                "旧版原因码": _readable_reason(row.get("V1原因码"), old_pass),
            }
        )
    return readable


def write_readable_outputs(
    rows: list[dict[str, Any]],
    output_stem: Path,
    dataset_entries: Iterable[dict[str, Any]],
) -> tuple[Path, Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output_stem.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_stem.with_suffix(".csv")
    xlsx_path = output_stem.with_suffix(".xlsx")
    readable_rows = build_readable_rows(rows, dataset_entries)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=READABLE_HEADERS)
        writer.writeheader()
        writer.writerows(readable_rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SN对比回执"
    sheet.append(READABLE_HEADERS)
    for item in readable_rows:
        sheet.append([item[header] for header in READABLE_HEADERS])

    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = {
        "A": 30,
        "B": 14,
        "C": 26,
        "D": 26,
        "E": 26,
        "F": 18,
        "G": 18,
        "H": 52,
        "I": 60,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(READABLE_HEADERS)):
        for cell in row_cells:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column_letter in {"A", "C", "D", "E"}:
                cell.number_format = "@"
                cell.value = "" if cell.value is None else str(cell.value)
        sheet.row_dimensions[row_cells[0].row].height = 28
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"
    table = Table(displayName="SNCompareReadable", ref=f"A1:I{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    actual_xlsx_path = _save_workbook_with_unlocked_fallback(workbook, xlsx_path)
    return csv_path, actual_xlsx_path


def _save_workbook_with_unlocked_fallback(workbook: Any, xlsx_path: Path) -> Path:
    try:
        workbook.save(xlsx_path)
        return xlsx_path
    except PermissionError:
        fallback = xlsx_path.with_name(f"{xlsx_path.stem}_unlocked{xlsx_path.suffix}")
        workbook.save(fallback)
        return fallback


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare Guobu SN V1 and V2 results by order ID")
    parser.add_argument("--v1-jsonl", required=True)
    parser.add_argument("--v2-jsonl", required=True)
    parser.add_argument("--out", required=True, help="Output path without extension")
    parser.add_argument("--dataset-json", help="Optional SN acceptance dataset JSON for fixed readable report")
    parser.add_argument("--readable-out", help="Optional readable report output path without extension")
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args()
    rows = compare_result_sets(load_jsonl(Path(args.v1_jsonl)), load_jsonl(Path(args.v2_jsonl)))
    csv_path, json_path = write_outputs(rows, Path(args.out))
    print(f"CSV: {csv_path}")
    print(f"JSON: {json_path}")
    if args.dataset_json:
        dataset_entries = json.loads(Path(args.dataset_json).read_text(encoding="utf-8-sig"))
        readable_stem = Path(args.readable_out) if args.readable_out else Path(f"{args.out}_readable")
        readable_csv, readable_xlsx = write_readable_outputs(rows, readable_stem, dataset_entries)
        print(f"READABLE_CSV: {readable_csv}")
        print(f"READABLE_XLSX: {readable_xlsx}")


if __name__ == "__main__":
    main()
