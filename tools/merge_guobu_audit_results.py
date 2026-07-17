
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NETWORK_MARKERS = (
    "timeouterror",
    "timed out",
    "modelconnectionerror",
    "connect failed",
    "winerror 10060",
    "http error 500",
)


def read_jsonl(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def network_failure(row: dict[str, Any]) -> bool:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("manual_reason", "manual_reason_cn", "strategy")
    ).lower()
    return any(marker in text for marker in NETWORK_MARKERS)


def manual(row: dict[str, Any]) -> bool:
    return bool(str(row.get("manual_reason") or row.get("manual_reason_cn") or "").strip())


def number(row: dict[str, Any], key: str) -> float:
    try:
        return float(row.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def integer(row: dict[str, Any], key: str) -> int:
    try:
        return int(row.get(key) or 0)
    except (TypeError, ValueError):
        return 0


def reason_text(row: dict[str, Any]) -> str:
    return str(row.get("manual_reason_cn") or row.get("manual_reason") or "")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--first-jsonl", required=True)
    parser.add_argument("--second-jsonl")
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--output-json", required=True)
    args = parser.parse_args()

    first_items = read_jsonl(Path(args.first_jsonl))
    second_items = read_jsonl(Path(args.second_jsonl)) if args.second_jsonl else []
    if not first_items:
        raise SystemExit("First-run JSONL is empty")

    second_by_id = {str(item["row"]["id"]): item for item in second_items}
    first_timeout_ids = {
        str(item["row"]["id"])
        for item in first_items
        if network_failure(item.get("row") or {})
    }

    rows: list[dict[str, Any]] = []
    for seq, first_item in enumerate(first_items, 1):
        first = first_item.get("row") or {}
        order_id = str(first.get("id") or first_item.get("task", {}).get("channel_order_no") or "")
        second_item = second_by_id.get(order_id) if order_id in first_timeout_ids else None
        final = (second_item or first_item).get("row") or {}
        first_elapsed = number(first, "elapsed_sec")
        second_elapsed = number(final, "elapsed_sec") if second_item else 0.0
        first_tokens = integer(first, "total_tokens")
        second_tokens = integer(final, "total_tokens") if second_item else 0
        rows.append(
            {
                "搴忓彿": seq,
                "璁㈠崟鍙?: order_id,
                "鏄惁杞汉宸?: "鏄? if manual(final) else "鍚?,
                "杞汉宸ュ師鍥犵爜": str(final.get("manual_reason_code") or ""),
                "杞汉宸ュ師鍥?: reason_text(final),
                "绯荤粺SN": str(final.get("system_sn") or ""),
                "妯″瀷SN": str(final.get("observed_sn") or ""),
                "鍘熸湰娴佺▼鐘舵€?: str(final.get("source_flow_status") or ""),
                "婧愬鏍哥姸鎬?: final.get("source_examine_status", ""),
                "婧愮粨绠楃姸鎬?: final.get("source_settle_status", ""),
                "鑰楁椂(绉?": round(number(final, "elapsed_sec"), 2),
                "浣跨敤token閲?: integer(final, "total_tokens"),
                "缁撴灉鏉ユ簮": "绗簩杞綉缁滈噸璺? if second_item else "绗竴娆″叏閲忓鏍?,
                "绗竴娆℃槸鍚︾綉缁滃け璐?: "鏄? if order_id in first_timeout_ids else "鍚?,
                "绗簩娆℃槸鍚︿粛缃戠粶澶辫触": "鏄? if second_item and network_failure(final) else ("鍚? if second_item else ""),
                "绗竴娆¤浆浜哄伐鍘熷洜": reason_text(first),
                "绗竴娆¤€楁椂(绉?": round(first_elapsed, 2),
                "绗簩娆¤€楁椂(绉?": round(second_elapsed, 2) if second_item else "",
                "涓よ疆鍚堣鑰楁椂(绉?": round(first_elapsed + second_elapsed, 2),
                "绗竴娆oken": first_tokens,
                "绗簩娆oken": second_tokens if second_item else "",
                "涓よ疆鍚堣token": first_tokens + second_tokens,
                "绛栫暐": str(final.get("strategy") or ""),
                "缃俊搴?: final.get("confidence", ""),
            }
        )

    final_manual = sum(row["鏄惁杞汉宸?] == "鏄? for row in rows)
    second_still_timeout = sum(item["row"]["id"] in first_timeout_ids and network_failure(item["row"]) for item in second_items)
    reason_counts = Counter(
        row["杞汉宸ュ師鍥犵爜"] if row["鏄惁杞汉宸?] == "鏄? and row["杞汉宸ュ師鍥犵爜"] else "鑷姩閫氳繃"
        for row in rows
    )
    summary = {
        "total": len(rows),
        "first_timeout": len(first_timeout_ids),
        "second_rerun": len(second_items),
        "second_resolved": len(second_items) - second_still_timeout,
        "second_still_timeout": second_still_timeout,
        "final_manual": final_manual,
        "final_auto": len(rows) - final_manual,
        "final_timeout": second_still_timeout,
        "final_tokens": sum(int(row["浣跨敤token閲?]) for row in rows),
        "actual_tokens": sum(int(row["涓よ疆鍚堣token"]) for row in rows),
        "final_elapsed_seconds": round(sum(float(row["鑰楁椂(绉?"]) for row in rows), 2),
        "actual_elapsed_seconds": round(sum(float(row["涓よ疆鍚堣鑰楁椂(绉?"]) for row in rows), 2),
    }

    output_xlsx = Path(args.output_xlsx)
    output_json = Path(args.output_json)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    detail = workbook.active
    detail.title = "鏄庣粏琛?
    headers = list(rows[0])
    detail.append(headers)
    for row in rows:
        detail.append([row[header] for header in headers])

    summary_sheet = workbook.create_sheet("姹囨€昏〃")
    summary_sheet.append(["鎸囨爣", "鏁板€?])
    summary_labels = {
        "total": "鏍锋湰鎬绘暟",
        "first_timeout": "绗竴娆＄綉缁滃け璐ヨ浆浜哄伐鍗曟暟",
        "second_rerun": "绗簩杞噸璺戝崟鏁?,
        "second_resolved": "绗簩杞凡瑙ｅ喅缃戠粶澶辫触鍗曟暟",
        "second_still_timeout": "绗簩杞粛缃戠粶澶辫触鍗曟暟",
        "final_manual": "缁煎悎鍚庤浆浜哄伐鎬绘暟",
        "final_auto": "缁煎悎鍚庤嚜鍔ㄩ€氳繃鏁?,
        "final_timeout": "缁煎悎鍚庣綉缁滃け璐ヨ浆浜哄伐鏁?,
        "final_tokens": "缁煎悎缁撴灉token鍚堣",
        "actual_tokens": "涓よ疆瀹為檯token鍚堣",
        "final_elapsed_seconds": "缁煎悎缁撴灉鑰楁椂鍚堣(绉?",
        "actual_elapsed_seconds": "涓よ疆瀹為檯鑰楁椂鍚堣(绉?",
    }
    for key, value in summary.items():
        summary_sheet.append([summary_labels[key], value])
    summary_sheet.append([])
    summary_sheet.append(["缁煎悎鍚庤浆浜哄伐鍘熷洜鍒嗗竷", "鍗曟暟"])
    for code, count in reason_counts.most_common():
        summary_sheet.append([code, count])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.font.bold, color=cell.font.color)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    widths = [8, 24, 12, 20, 55, 28, 28, 16, 12, 12, 12, 14, 18, 16, 18, 55, 16, 16, 18, 14, 14, 16, 24, 10]
    for index, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(index)].width = width
    summary_sheet.column_dimensions["A"].width = 36
    summary_sheet.column_dimensions["B"].width = 18

    table = Table(displayName="GuobuAuditDetail", ref=f"A1:{get_column_letter(detail.max_column)}{detail.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    detail.add_table(table)
    workbook.save(output_xlsx)

    output_json.write_text(
        json.dumps({"summary": summary, "reason_counts": dict(reason_counts)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    check = load_workbook(output_xlsx, read_only=True)
    if check["鏄庣粏琛?].max_row != len(rows) + 1:
        raise SystemExit("Workbook detail row count mismatch")
    if any("??" in value for sheet in check.worksheets for row in sheet.iter_rows(values_only=True) for value in row if isinstance(value, str)):
        raise SystemExit("Workbook contains question-mark encoding damage")

    print(json.dumps({"xlsx": str(output_xlsx), "json": str(output_json), "summary": summary}, ensure_ascii=False))


if __name__ == "__main__":
    main()
