import json
import subprocess
import sys

import pytest
from openpyxl import load_workbook

from tools.guobu_audit_report import (
    _retry_ids,
    build_summary,
    merge_attempts,
    network_failure,
    parse_manual_flag,
    sn_display,
    standard_reason,
    write_report,
)


REASONS = {
    "PRODUCT_TYPE_MISMATCH": "商品类型不一致",
    "PRODUCT_PHOTO_INVALID": "商品照片不符合要求",
    "UNBOXING_PHOTO_INVALID": "拆封/安装照片不符合要求",
    "ACTIVATION_PHOTO_INVALID": "激活照片不符合要求",
    "SN_MISSING_IN_ACTIVATION_PHOTO": "激活照片不符合要求",
    "ADDRESS_TOO_COARSE": "收货地址不符合要求",
    "DUPLICATE_IMAGE_EVIDENCE": "存在重复图片，不符合要求",
    "NON_REAL_PHOTO_REVIEW": "图片疑似非实拍",
    "NON_REAL_PHOTO_STRONG_RISK": "图片疑似非实拍",
    "IMAGE_STRONG_RISK": "图片疑似非实拍",
    "SN_MISMATCH": "SN不一致",
    "INVOICE_ORANGE_WARNING": "发票疑似已红冲",
    "MODEL_UNCERTAIN": "图片信息无法确认",
    "PHOTO_AUTHENTICITY_SERVICE_FAILURE": "审核服务异常",
    "ARTIFACT_LOAD_FAILURE": "审核服务异常",
    "FFT_FAILURE": "审核服务异常",
    "SN_TRUNCATED_OBSCURED": "SN不完整，无法识别",
    "SN_NOT_FOUND": "SN无法识别",
    "SYSTEM_SN_MISSING": "系统SN缺失",
    "IMAGE_MISSING": "图片缺失",
    "FIELD_MISSING": "订单信息缺失",
    "PRODUCT_TYPE_MISSING": "商品类型信息缺失",
    "NON_REAL_PHOTO_FFT_RESCUE": "图片疑似非实拍",
}


@pytest.mark.parametrize(("code", "text"), REASONS.items())
def test_standard_reason_maps_every_confirmed_code(code, text):
    assert standard_reason(code) == text


def test_standard_reason_handles_empty_and_unknown_codes():
    assert standard_reason("") == ""
    assert standard_reason("NEW_REASON") == "图片信息无法确认"


@pytest.mark.parametrize(
    ("value", "expected"),
    [(True, True), (False, False), ("是", True), ("否", False)],
)
def test_parse_manual_flag_accepts_only_explicit_values(value, expected):
    assert parse_manual_flag(value) is expected


@pytest.mark.parametrize(
    "value",
    [None, 0, 1, "", "true", " 是", "否 ", "\u93c4\u75d0", "\u935a\ue6c6"],
)
def test_parse_manual_flag_rejects_ambiguous_or_corrupt_values(value):
    with pytest.raises(ValueError):
        parse_manual_flag(value)


@pytest.mark.parametrize(
    ("row", "expected"),
    [
        ({"sn_match": True, "system_sn": "", "observed_sn": ""}, ("是", "")),
        ({"sn_match": False, "system_sn": "", "observed_sn": "ABC"}, ("无系统SN", "")),
        (
            {"sn_match": False, "system_sn": "ABC", "observed_sn": ""},
            ("未读取", "模型未读取到SN"),
        ),
        ({"sn_match": False, "system_sn": "ABC", "observed_sn": "ABC"}, ("否", "")),
    ],
)
def test_sn_display_uses_fixed_status_priority(row, expected):
    assert sn_display(row) == expected


@pytest.mark.parametrize(
    ("system_sn", "observed_sn", "difference"),
    [
        ("ABOCD", "AB0CD", "第3位不同：系统O，模型0"),
        ("ABCD", "SABCD", "模型第1位多读S"),
        ("ABCD", "ABXCD", "模型第3位多读X"),
        ("ABXCD", "ABCD", "模型第3位少读X"),
        ("ABCD", "ACBD", "字符顺序不同：系统BC，模型CB"),
        ("ABCDJ", "ABCD", "模型末尾少读J"),
        ("PQABCD", "ABCD", "模型开头少读PQ"),
        ("ABCD", "AB", "模型末尾少读CD"),
        ("ABCD", "CD", "模型开头少读AB"),
        ("00123", "00123", ""),
        ("ABCD", "AXYD", "SN存在多处差异"),
        ("ABCDE", "AXC", "SN存在多处差异"),
    ],
)
def test_sn_display_classifies_raw_string_differences(system_sn, observed_sn, difference):
    row = {"sn_match": False, "system_sn": system_sn, "observed_sn": observed_sn}
    assert sn_display(row) == ("否", difference)


def test_sn_display_does_not_mutate_sn_match_or_apply_visual_equivalence():
    row = {"sn_match": False, "system_sn": "O123", "observed_sn": "0123"}
    assert sn_display(row) == ("否", "第1位不同：系统O，模型0")
    assert row["sn_match"] is False


def audit_item(order_id="1", *, flag=False, code="", reason="", status="\u5df2\u901a\u8fc7", elapsed=1.0, raw=None, error=""):
    item = {"row": {"id": order_id, "manual_flag": flag, "manual_reason_code": code,
                    "manual_reason": reason, "source_flow_status": status,
                    "elapsed_sec": elapsed, "_raw": raw or {}}}
    if error:
        item["_error"] = error
    return item


@pytest.mark.parametrize("order_id", ["", None])
def test_empty_first_id_fails_closed(order_id):
    with pytest.raises(ValueError, match="first-run order ID"):
        merge_attempts([audit_item(order_id)], [])


def test_duplicate_first_id_fails_closed():
    with pytest.raises(ValueError, match="duplicate first-run order ID"):
        merge_attempts([audit_item("1"), audit_item("1")], [])


def test_duplicate_retry_id_fails_closed():
    first = audit_item("1", error="TimeoutError")
    with pytest.raises(ValueError, match="duplicate retry order ID"):
        merge_attempts([first], [audit_item("1"), audit_item("1")])


def test_unknown_retry_id_fails_closed():
    with pytest.raises(ValueError, match="unknown retry order ID"):
        merge_attempts([audit_item("1")], [audit_item("2")])


def test_network_failure_reads_item_error_and_row_fields_case_insensitively():
    assert network_failure(audit_item(error="ModelConnectionError: unavailable"))
    assert network_failure(audit_item(reason="request TIMED OUT"))
    assert network_failure({"row": {"manual_reason_cn": "Connect Failed"}})
    assert network_failure({"row": {"strategy": "HTTP ERROR 500 fallback"}})
    assert not network_failure(audit_item(reason="SN mismatch"))


def test_retry_selection_must_match_items_and_detected_failures():
    first = [audit_item("1", error="TimeoutError"), audit_item("2", error="WinError 10060")]
    with pytest.raises(ValueError, match="retry selection mismatch"):
        merge_attempts(first, [audit_item("1")], retry_ids={"1", "2"})
    with pytest.raises(ValueError, match="not a first-run network failure"):
        merge_attempts([audit_item("1")], [audit_item("1")], retry_ids={"1"})


def test_cli_accepts_completed_retry_selection_object(tmp_path):
    first_item = audit_item("9001", flag=True, code="MODEL_UNCERTAIN",
                            error="TimeoutError")
    retry_item = audit_item("9001", flag=False)
    for item in (first_item, retry_item):
        item["row"].update(system_sn="0001", observed_sn="0001", sn_match=True)
    first, retry = tmp_path / "first.jsonl", tmp_path / "retry.jsonl"
    first.write_text(json.dumps(first_item) + "\n", encoding="utf-8")
    retry.write_text(json.dumps(retry_item) + "\n", encoding="utf-8")
    selection = tmp_path / "selection.json"
    selection.write_text(json.dumps({"source_dirs": ["completed"], "requested": 1,
        "selected": 1, "missing": [], "orders": ["9001"], "out_dir": "output"}),
        encoding="utf-8")
    xlsx, output_json = tmp_path / "cli.xlsx", tmp_path / "cli.json"
    script = str((__import__("pathlib").Path(__file__).parents[1]
                  / "tools" / "guobu_audit_report.py"))
    completed = subprocess.run([sys.executable, script, "--first-jsonl", str(first),
        "--retry-jsonl", str(retry), "--retry-selection-json", str(selection),
        "--output-xlsx", str(xlsx), "--output-json", str(output_json)],
        text=True, capture_output=True, timeout=30)
    assert completed.returncode == 0, completed.stderr
    assert json.loads(output_json.read_text(encoding="utf-8"))["rows"][0]["final_source"] == "retry"


@pytest.mark.parametrize("selection", [
    {"requested": 1, "selected": 1, "missing": []},
    {"requested": 1, "selected": 1, "missing": [], "orders": "9001"},
    {"requested": 2, "selected": 1, "missing": [], "orders": ["9001"]},
    {"requested": 1, "selected": 1, "missing": ["9002"], "orders": ["9001"]},
])
def test_retry_selection_object_fails_closed_on_malformed_metadata(selection, tmp_path):
    path = tmp_path / "selection.json"
    path.write_text(json.dumps(selection), encoding="utf-8")
    with pytest.raises(ValueError, match="retry selection"):
        _retry_ids(str(path))


def test_retry_selection_retains_bare_list_compatibility(tmp_path):
    path = tmp_path / "selection.json"
    path.write_text(json.dumps(["9001", 9002]), encoding="utf-8")
    assert _retry_ids(str(path)) == {"9001", "9002"}


@pytest.mark.parametrize("orders", [
    ["1", "1"],
    ["1", " 1 "],
])
def test_retry_selection_object_rejects_normalized_duplicate_ids(orders, tmp_path):
    path = tmp_path / "selection.json"
    path.write_text(json.dumps({"requested": 2, "selected": 2, "missing": [],
                                "orders": orders}), encoding="utf-8")
    with pytest.raises(ValueError, match="duplicate retry selection order ID"):
        _retry_ids(str(path))


@pytest.mark.parametrize("bad_id", ["", "   ", True, None, ["1"], {"id": "1"}])
def test_retry_selection_object_rejects_empty_or_non_scalar_ids(bad_id, tmp_path):
    path = tmp_path / "selection.json"
    path.write_text(json.dumps({"requested": 1, "selected": 1, "missing": [],
                                "orders": [bad_id]}), encoding="utf-8")
    with pytest.raises(ValueError, match="retry selection order ID"):
        _retry_ids(str(path))


@pytest.mark.parametrize("orders", [["1", " 1 "], [""], [False], [["1"]]])
def test_bare_retry_selection_list_uses_same_id_validation(orders, tmp_path):
    path = tmp_path / "selection.json"
    path.write_text(json.dumps(orders), encoding="utf-8")
    with pytest.raises(ValueError, match="retry selection order ID"):
        _retry_ids(str(path))


@pytest.mark.parametrize("item", [
    audit_item(flag="maybe"),
    audit_item(flag=False, code="SN_MISMATCH"),
    audit_item(flag=False, reason="business failure"),
    audit_item(flag=True, code=""),
])
def test_final_flag_and_primary_reason_contradictions_fail_closed(item):
    with pytest.raises(ValueError):
        merge_attempts([item], [])


def test_retry_replaces_legal_failure_and_retains_both_attempts():
    first = audit_item("1", flag=True, code="MODEL_UNCERTAIN", error="TimeoutError", elapsed=30)
    retry = audit_item("1", flag=False, elapsed=2)
    rows, accounting = merge_attempts([first], [retry])
    assert rows[0]["row"]["manual_flag"] is False
    assert rows[0]["first_attempt"] is first
    assert rows[0]["retry_attempt"] is retry
    assert [attempt["source"] for attempt in accounting["attempts"]] == ["first", "retry"]
    assert accounting["elapsed_seconds"] == 32


def test_network_failure_elapsed_is_capped_but_raw_trace_is_retained():
    failed = audit_item("1", flag=True, code="MODEL_UNCERTAIN",
                        error="TimeoutError", elapsed=10811.89)
    retry = audit_item("1", flag=False, elapsed=2.5)
    normal = audit_item("2", elapsed=10)
    _, accounting = merge_attempts([failed, normal], [retry])
    assert accounting["raw_elapsed_seconds"] == pytest.approx(10824.39)
    assert accounting["elapsed_seconds"] == pytest.approx(72.5)
    assert accounting["order_timeout_seconds"] == 60
    assert [(attempt["elapsed_seconds"], attempt["effective_elapsed_seconds"])
            for attempt in accounting["attempts"]] == [
                (10811.89, 60), (10, 10), (2.5, 2.5)]


def test_network_failure_elapsed_uses_configured_timeout_and_nonnegative_floor():
    failures = [
        audit_item("1", flag=True, code="MODEL_UNCERTAIN",
                   error="TimeoutError", elapsed=40500.92),
        audit_item("2", flag=True, code="MODEL_UNCERTAIN",
                   error="TimeoutError", elapsed=-4),
    ]
    _, accounting = merge_attempts(failures, [], order_timeout_seconds=30)
    assert accounting["raw_elapsed_seconds"] == pytest.approx(40496.92)
    assert accounting["elapsed_seconds"] == 30
    assert accounting["order_timeout_seconds"] == 30


@pytest.mark.parametrize("timeout", [0, -1, float("inf"), float("nan"), True])
def test_order_timeout_must_be_positive_finite_number(timeout):
    with pytest.raises(ValueError, match="order timeout"):
        merge_attempts([audit_item()], [], order_timeout_seconds=timeout)


def test_valid_retry_cannot_bypass_malformed_first_attempt_flag():
    first = audit_item("1", flag="invalid", error="TimeoutError")
    retry = audit_item("1", flag=False)
    with pytest.raises(ValueError, match="invalid manual_flag"):
        merge_attempts([first], [retry])


def test_only_final_primary_reason_code_is_exposed():
    item = audit_item(flag=True, code="SN_MISMATCH", reason="mentions IMAGE_MISSING")
    item["row"]["manual_reason_codes"] = ["SN_MISMATCH", "IMAGE_MISSING"]
    rows, _ = merge_attempts([item], [])
    assert rows[0]["reason_code"] == "SN_MISMATCH"
    assert rows[0]["reason"] == standard_reason("SN_MISMATCH")


def test_cached_stage_usage_is_logical_but_not_billed_and_keys_are_unique():
    usage = {"prompt_tokens": 100, "completion_tokens": 20}
    raw = {"sn_usage": usage, "sn_cached": True,
           "review_usage": {"prompt_tokens": 50, "completion_tokens": 10},
           "review_cached": False, "review_usage_alias": usage}
    _, accounting = merge_attempts([audit_item(raw=raw)], [])
    assert accounting["logical_input_tokens"] == 150
    assert accounting["logical_output_tokens"] == 30
    assert accounting["billed_input_tokens"] == 50
    assert accounting["billed_output_tokens"] == 10


def test_cached_input_tokens_are_separated_and_priced_at_cached_rate():
    raw = {"compliance_usage": {"prompt_tokens": 100,
            "prompt_tokens_details": {"cached_tokens": 40}, "completion_tokens": 10},
           "compliance_cached": False}
    rows, accounting = merge_attempts([audit_item(raw=raw)], [])
    summary = build_summary(rows, accounting, {"input_per_million": 2,
        "cached_input_per_million": 1, "output_per_million": 3})
    assert accounting["billed_input_tokens"] == 60
    assert accounting["billed_cached_input_tokens"] == 40
    assert accounting["logical_input_tokens"] == 60
    assert accounting["logical_cached_input_tokens"] == 40
    assert summary["estimated_cost"] == pytest.approx(0.00019)
    assert build_summary(rows, accounting)["estimated_cost"] == "\u5f85\u914d\u7f6e"


def test_same_raw_usage_root_is_not_counted_twice():
    raw = {"sn_usage": {"prompt_tokens": 10, "completion_tokens": 2}, "sn_cached": False}
    item = audit_item(raw=raw)
    item["_raw"] = raw
    _, accounting = merge_attempts([item], [])
    assert accounting["billed_input_tokens"] == 10
    assert accounting["billed_output_tokens"] == 2


def test_same_usage_object_across_distinct_raw_roots_is_counted_once():
    usage = {"prompt_tokens": 10, "completion_tokens": 2}
    item = audit_item(raw={"sn_usage": usage, "sn_cached": False})
    item["_raw"] = {"review_usage": usage, "review_cached": False}
    _, accounting = merge_attempts([item], [])
    assert accounting["logical_input_tokens"] == 10
    assert accounting["logical_output_tokens"] == 2
    assert accounting["billed_input_tokens"] == 10
    assert accounting["billed_output_tokens"] == 2


def test_equal_but_distinct_usage_objects_are_each_counted():
    item = audit_item(raw={"sn_usage": {"prompt_tokens": 10, "completion_tokens": 2},
                           "sn_cached": False})
    item["_raw"] = {"review_usage": {"prompt_tokens": 10, "completion_tokens": 2},
                    "review_cached": False}
    _, accounting = merge_attempts([item], [])
    assert accounting["logical_input_tokens"] == 20
    assert accounting["logical_output_tokens"] == 4
    assert accounting["billed_input_tokens"] == 20
    assert accounting["billed_output_tokens"] == 4


def test_summary_uses_exact_trimmed_statuses_and_excludes_unknowns():
    rows, accounting = merge_attempts([
        audit_item("1", flag=True, code="SN_MISMATCH", status="  \u672a\u901a\u8fc7 "),
        audit_item("2", status="\u5df2\u901a\u8fc7"),
        audit_item("3", flag=True, code="SN_MISMATCH", status="pending"),
        audit_item("4", flag=True, code="SN_MISMATCH", status="\u5df2\u901a\u8fc7-extra"),
    ], [])
    summary = build_summary(rows, accounting)
    assert summary["failed_interception"] == {"numerator": 1, "denominator": 1, "rate": 1.0}
    assert summary["passed_false_positive"] == {"numerator": 0, "denominator": 1, "rate": 0.0}


def test_zero_denominators_have_no_rate():
    rows, accounting = merge_attempts([audit_item(status="reviewing")], [])
    summary = build_summary(rows, accounting)
    assert summary["failed_interception"] == {"numerator": 0, "denominator": 0, "rate": None}
    assert summary["passed_false_positive"] == {"numerator": 0, "denominator": 0, "rate": None}


def test_efficiency_uses_baseline_and_accumulated_elapsed():
    rows, accounting = merge_attempts([audit_item("1", elapsed=1800), audit_item("2", elapsed=1800)], [])
    summary = build_summary(rows, accounting)
    baseline = 550 / 7.5
    assert summary["effective_hours"] == 1
    assert summary["human_orders_per_hour"] == pytest.approx(baseline)
    assert summary["human_estimated_hours"] == pytest.approx(2 / baseline)
    assert summary["model_orders_per_hour"] == 2
    assert summary["efficiency_multiple"] == pytest.approx(2 / baseline)
    assert summary["efficiency_improvement"] == pytest.approx(2 / baseline - 1)
    assert summary["saved_human_hours"] == pytest.approx(2 / baseline - 1)


def test_zero_elapsed_avoids_division_by_zero():
    rows, accounting = merge_attempts([audit_item(elapsed=0)], [])
    summary = build_summary(rows, accounting)
    assert summary["model_orders_per_hour"] is None
    assert summary["efficiency_multiple"] is None
    assert summary["efficiency_improvement"] is None


def test_sn_display_transposition_names_the_raw_character_pairs():
    row = {"sn_match": False, "system_sn": "12XY34", "observed_sn": "12YX34"}
    assert sn_display(row) == ("否", "字符顺序不同：系统XY，模型YX")


def report_fixture(order_id="001234567890123456789", system_sn="00123", observed_sn="00124"):
    item = audit_item(order_id, flag=True, code="SN_MISMATCH", reason="=external",
                      status="未通过", elapsed=3600,
                      raw={"review_usage": {"prompt_tokens": 100, "completion_tokens": 20}})
    item["row"].update(system_sn=system_sn, observed_sn=observed_sn, sn_match=False)
    rows, accounting = merge_attempts([item], [])
    prices = {"input_per_million": 2, "cached_input_per_million": 1,
              "output_per_million": 3}
    summary = build_summary(rows, accounting, prices)
    return rows, summary, {"summary": summary, "accounting": accounting,
                           "pricing": prices, "rows": rows}


def test_workbook_has_exact_business_contract_and_safe_text(tmp_path):
    rows, summary, audit_json = report_fixture()
    xlsx, output_json = tmp_path / "report.xlsx", tmp_path / "report.json"
    write_report(rows, summary, audit_json, xlsx, output_json)
    wb = load_workbook(xlsx, data_only=False)
    assert wb.sheetnames == ["明细表", "汇总表"]
    detail = wb["明细表"]
    assert [cell.value for cell in detail[1]] == [
        "订单号", "是否转人工", "原始流程状态", "转人工原因",
        "系统SN", "模型SN", "SN是否一致", "SN具体差别",
    ]
    assert detail.max_row == 2
    assert [detail.cell(2, col).value for col in (1, 5, 6)] == [
        "001234567890123456789", "00123", "00124"]
    assert all(detail.cell(2, col).number_format == "@" for col in (1, 5, 6))
    assert detail["D2"].data_type == "s" and not str(detail["D2"].value).startswith("=")
    assert detail.freeze_panes == "A2"
    assert detail.auto_filter.ref == detail.tables["明细表格"].ref
    assert detail["D2"].alignment.wrap_text and detail["H2"].alignment.wrap_text
    assert all(cell.font.name == "Arial" for sheet in wb for row in sheet for cell in row)
    assert detail["A1"].font.bold and detail["A1"].font.color.rgb.endswith("FFFFFF")

    sheet = wb["汇总表"]
    metrics = {sheet.cell(row, 1).value: sheet.cell(row, 2)
               for row in range(2, sheet.max_row + 1)}
    assert metrics["未通过拦截率"].value.startswith("=IF(")
    assert metrics["未通过拦截率"].number_format == "0.0%"
    assert metrics["有效审核总用时（小时）"].number_format == "0.00"
    assert metrics["效率倍数"].number_format == "0.0x"
    assert "累计每订单处理时长" in metrics["口径说明"].value


def test_workbook_elapsed_note_includes_timeout_and_raw_effective_totals(tmp_path):
    rows, summary, audit_json = report_fixture()
    xlsx = tmp_path / "elapsed-note.xlsx"
    write_report(rows, summary, audit_json, xlsx, tmp_path / "elapsed-note.json")
    note = load_workbook(xlsx, data_only=False)["\u6c47\u603b\u8868"]["B22"].value
    assert "60" in note
    assert note.count("3600.00") == 2


def test_workbook_zero_denominator_and_unconfigured_cost(tmp_path):
    rows, accounting = merge_attempts([audit_item(status="reviewing")], [])
    summary = build_summary(rows, accounting)
    xlsx = tmp_path / "zero.xlsx"
    write_report(rows, summary, {"summary": summary, "accounting": accounting, "pricing": None},
                 xlsx, tmp_path / "zero.json")
    sheet = load_workbook(xlsx, data_only=False)["汇总表"]
    metrics = {sheet.cell(row, 1).value: sheet.cell(row, 2).value
               for row in range(2, sheet.max_row + 1)}
    assert 'IF(B' in metrics["未通过拦截率"] and '"无可计算样本"' in metrics["未通过拦截率"]
    assert metrics["Token预计成本"] == "待配置"


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@"])
def test_formula_injection_strings_round_trip_as_text(prefix, tmp_path):
    rows, summary, audit_json = report_fixture(system_sn=prefix + "SN", observed_sn=prefix + "MODEL")
    rows[0]["row"]["source_flow_status"] = prefix + "STATUS"
    xlsx = tmp_path / f"safe-{ord(prefix)}.xlsx"
    write_report(rows, summary, audit_json, xlsx, tmp_path / f"safe-{ord(prefix)}.json")
    detail = load_workbook(xlsx, data_only=False)["明细表"]
    assert [(detail.cell(2, col).value, detail.cell(2, col).data_type) for col in (3, 5, 6)] == [
        (prefix + "STATUS", "s"), (prefix + "SN", "s"), (prefix + "MODEL", "s")]


def test_json_utf8_trace_and_validation(tmp_path):
    rows, summary, audit_json = report_fixture()
    xlsx, output_json = tmp_path / "trace.xlsx", tmp_path / "trace.json"
    write_report(rows, summary, audit_json, xlsx, output_json)
    payload = json.loads(output_json.read_text(encoding="utf-8"))
    assert payload["rows"][0]["reason_code"] == "SN_MISMATCH"
    assert payload["rows"][0]["row"]["system_sn"] == "00123"
    assert payload["accounting"]["attempts"][0]["source"] == "first"
    assert payload["pricing"]["input_per_million"] == 2
    with pytest.raises(FileExistsError):
        write_report(rows, summary, audit_json, xlsx, output_json)
    with pytest.raises(ValueError, match="empty rows"):
        write_report([], summary, audit_json, tmp_path / "empty.xlsx", tmp_path / "empty.json")
    with pytest.raises(ValueError, match="duplicate display order ID"):
        write_report(rows + rows, summary, audit_json, tmp_path / "dup.xlsx", tmp_path / "dup.json")


def test_cli_regenerates_from_jsonl_without_network_access(tmp_path):
    item = audit_item("9001", flag=False, elapsed=2)
    item["row"].update(system_sn="0001", observed_sn="0001", sn_match=True)
    first = tmp_path / "first.jsonl"
    first.write_text(json.dumps(item, ensure_ascii=False) + "\n", encoding="utf-8")
    xlsx, output_json = tmp_path / "cli.xlsx", tmp_path / "cli.json"
    script = str((__import__("pathlib").Path(__file__).parents[1] / "tools" / "guobu_audit_report.py"))
    completed = subprocess.run([sys.executable, script, "--first-jsonl", str(first),
        "--output-xlsx", str(xlsx), "--output-json", str(output_json)],
        text=True, capture_output=True, timeout=30)
    assert completed.returncode == 0, completed.stderr
    assert xlsx.exists() and output_json.exists()
    assert load_workbook(xlsx)["明细表"]["A2"].value == "9001"


def test_cli_applies_configured_network_failure_timeout(tmp_path):
    item = audit_item("9001", flag=True, code="MODEL_UNCERTAIN",
                      error="TimeoutError", elapsed=10811.89)
    item["row"].update(system_sn="0001", observed_sn="", sn_match=False)
    first = tmp_path / "first.jsonl"
    first.write_text(json.dumps(item) + "\n", encoding="utf-8")
    xlsx, output_json = tmp_path / "cli.xlsx", tmp_path / "cli.json"
    script = str((__import__("pathlib").Path(__file__).parents[1]
                  / "tools" / "guobu_audit_report.py"))
    completed = subprocess.run([sys.executable, script, "--first-jsonl", str(first),
        "--order-timeout-seconds", "30", "--output-xlsx", str(xlsx),
        "--output-json", str(output_json)], text=True, capture_output=True, timeout=30)
    assert completed.returncode == 0, completed.stderr
    accounting = json.loads(output_json.read_text(encoding="utf-8"))["accounting"]
    assert accounting["raw_elapsed_seconds"] == 10811.89
    assert accounting["elapsed_seconds"] == 30
    assert accounting["order_timeout_seconds"] == 30


@pytest.mark.parametrize("bad_item", [
    audit_item("1", flag=True, code=""),
    audit_item("1", flag=False, code="SN_MISMATCH"),
])
def test_write_report_revalidates_final_business_decision(bad_item, tmp_path):
    row = bad_item["row"]
    merged = {"row": row, "manual": bool(row["manual_flag"]),
              "reason_code": str(row["manual_reason_code"]), "reason": "caller supplied",
              "final_source": "first"}
    _, _, audit_json = report_fixture()
    with pytest.raises(ValueError):
        write_report([merged], audit_json["summary"], audit_json,
                     tmp_path / "bad.xlsx", tmp_path / "bad.json")


def test_write_report_derives_reason_from_primary_code_not_caller(tmp_path):
    rows, summary, audit_json = report_fixture()
    rows[0]["reason"] = "=untrusted caller reason"
    xlsx = tmp_path / "derived.xlsx"
    write_report(rows, summary, audit_json, xlsx, tmp_path / "derived.json")
    assert load_workbook(xlsx)["明细表"]["D2"].value == standard_reason("SN_MISMATCH")


@pytest.mark.parametrize("audit_json", [
    {},
    {"accounting": {}, "pricing": None},
    {"accounting": {"attempts": [{}]}, "pricing": None},
    {"accounting": {"attempts": []}, "pricing": {"input_per_million": "bad"}},
])
def test_write_report_requires_validated_audit_trace(audit_json, tmp_path):
    rows, summary, _ = report_fixture()
    with pytest.raises(ValueError, match="audit"):
        write_report(rows, summary, audit_json, tmp_path / "audit.xlsx", tmp_path / "audit.json")


def test_write_report_requires_explicit_pricing_assumption(tmp_path):
    rows, summary, audit_json = report_fixture()
    del audit_json["pricing"]
    with pytest.raises(ValueError, match="audit pricing"):
        write_report(rows, summary, audit_json, tmp_path / "pricing.xlsx", tmp_path / "pricing.json")


@pytest.mark.parametrize("timeout", [0, -1, float("inf"), float("nan"), True])
def test_write_report_rejects_invalid_order_timeout_assumption(timeout, tmp_path):
    rows, summary, audit_json = report_fixture()
    audit_json["accounting"]["order_timeout_seconds"] = timeout
    with pytest.raises(ValueError, match="order timeout"):
        write_report(rows, summary, audit_json,
                     tmp_path / "timeout.xlsx", tmp_path / "timeout.json")


def test_write_report_replaces_stale_payload_summary_and_rows(tmp_path):
    rows, summary, audit_json = report_fixture()
    audit_json["summary"] = {"stale": True}
    audit_json["rows"] = [{"stale": True}]
    write_report(rows, summary, audit_json, tmp_path / "actual.xlsx", tmp_path / "actual.json")
    payload = json.loads((tmp_path / "actual.json").read_text(encoding="utf-8"))
    assert payload["summary"] == summary
    assert payload["rows"][0]["reason_code"] == "SN_MISMATCH"


def test_write_report_rejects_same_resolved_output_path(tmp_path):
    rows, summary, audit_json = report_fixture()
    output = tmp_path / "same"
    with pytest.raises(ValueError, match="distinct"):
        write_report(rows, summary, audit_json, output, output)


@pytest.mark.parametrize("source_lines", [
    ["not-json"],
    [json.dumps(audit_item("1")), json.dumps(audit_item("1"))],
])
def test_cli_overwrite_preserves_old_outputs_when_source_invalid(source_lines, tmp_path):
    first = tmp_path / "invalid.jsonl"
    first.write_text("\n".join(source_lines) + "\n", encoding="utf-8")
    xlsx, output_json = tmp_path / "old.xlsx", tmp_path / "old.json"
    xlsx.write_bytes(b"known-good-xlsx")
    output_json.write_text("known-good-json", encoding="utf-8")
    script = str((__import__("pathlib").Path(__file__).parents[1] / "tools" / "guobu_audit_report.py"))
    completed = subprocess.run([sys.executable, script, "--first-jsonl", str(first),
        "--output-xlsx", str(xlsx), "--output-json", str(output_json), "--overwrite"],
        text=True, capture_output=True, timeout=30)
    assert completed.returncode != 0
    assert xlsx.read_bytes() == b"known-good-xlsx"
    assert output_json.read_text(encoding="utf-8") == "known-good-json"
    assert not list(tmp_path.glob("*.tmp"))


def test_cli_overwrite_rejects_same_resolved_final_path_and_preserves_file(tmp_path):
    item = audit_item("1")
    item["row"].update(system_sn="1", observed_sn="1", sn_match=True)
    first = tmp_path / "first.jsonl"
    first.write_text(json.dumps(item) + "\n", encoding="utf-8")
    output = tmp_path / "good-output"
    output.write_bytes(b"known-good-output")
    script = str((__import__("pathlib").Path(__file__).parents[1] / "tools" / "guobu_audit_report.py"))
    completed = subprocess.run([sys.executable, script, "--first-jsonl", str(first),
        "--output-xlsx", str(output), "--output-json", str(output), "--overwrite"],
        text=True, capture_output=True, timeout=30)
    assert completed.returncode != 0
    assert output.read_bytes() == b"known-good-output"
    assert not list(tmp_path.glob("*.tmp.*"))
